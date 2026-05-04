"""Парсинг Excel: все листы → структура (если уверены) или LLM-фоллбэк."""

from __future__ import annotations

import asyncio
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.llm.client import LLMClient, LLMParseError
from core.parsers.types import ParseResult, ParsedItem

# Заголовки описания позиции — совпадение по подстроке (название колонки часто длинное).
DESC_KEYS = ("описан", "наимен", "позиц", "издел", "товар", "продукц", "артикул")

# Заголовки количества — совпадение ПО НАЧАЛУ слова, чтобы не цепляться за «количество вариантов».
QTY_KEYS = ("количеств", "кол-во", "кол.", "к-во", "qty", "объем", "объём")

# Если в названии колонки рядом со словом «количество/объём» есть это — НЕ qty-колонка.
QTY_DISQUALIFIERS = ("вариант", "дн", "посещ", "часов", "людей", "позиций", "сотрудник")

# Заголовки единицы измерения.
UNIT_KEYS = ("ед.изм", "ед. изм", "единиц", "ед-ц", "unit", "ед.")

# Число в свободном виде: «27.34», «3 916,99», «1 234.5», «3».
NUMBER_RE = re.compile(
    r"-?\d{1,3}(?:[  ]\d{3})+(?:[.,]\d+)?|-?\d+(?:[.,]\d+)?"
)


def _norm(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _is_total_row(s: str) -> bool:
    t = s.lower()
    return any(x in t for x in ("∑", "итого", "total", "всего", "сумма по"))


def _parse_qty_with_unit(cell: Any) -> tuple[Decimal | None, str | None]:
    """Извлекает (qty, unit) из ячейки. Поддерживает «27.34 м», «3 изделия», «1 шт - смонтированный»."""
    if cell is None:
        return None, None
    if isinstance(cell, bool):
        return None, None
    if isinstance(cell, (int, float)):
        try:
            return Decimal(str(cell)), None
        except InvalidOperation:
            return None, None
    s = str(cell).strip().replace(" ", " ")
    if not s:
        return None, None
    m = NUMBER_RE.search(s)
    if not m:
        return None, None
    num_str = m.group(0).replace(" ", "").replace(",", ".")
    try:
        qty = Decimal(num_str)
    except InvalidOperation:
        return None, None
    rest = (s[: m.start()] + s[m.end():]).strip(" -—–.,;:()/")
    unit = rest.split()[0] if rest else None
    if unit and len(unit) > 30:
        unit = None
    return qty, unit


class ExcelParser:
    def _score_header_row(
        self, row: tuple[Any, ...]
    ) -> tuple[int | None, int | None, int | None, int]:
        col_desc = col_qty = col_unit = None
        for j, raw in enumerate(row):
            cell = _norm(raw)
            if not cell:
                continue
            if col_qty is None and any(cell.startswith(k) for k in QTY_KEYS):
                if not any(d in cell for d in QTY_DISQUALIFIERS):
                    col_qty = j
            if col_desc is None and any(k in cell for k in DESC_KEYS):
                col_desc = j
            if col_unit is None and any(k in cell for k in UNIT_KEYS):
                col_unit = j
        score = sum(x is not None for x in (col_desc, col_qty, col_unit))
        return col_desc, col_qty, col_unit, score

    def _find_header(
        self, ws: Any
    ) -> tuple[int | None, int | None, int | None, int | None]:
        max_scan = min(ws.max_row or 30, 30)
        best: tuple[int, int | None, int | None, int | None, int] | None = None
        for ri, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1
        ):
            if not row:
                continue
            col_desc, col_qty, col_unit, score = self._score_header_row(row)
            if col_desc is None:
                continue
            if best is None or score > best[4]:
                best = (ri, col_desc, col_qty, col_unit, score)
        if best is None:
            return None, None, None, None
        return best[0], best[1], best[2], best[3]

    def _parse_sheet(self, ws: Any) -> tuple[list[ParsedItem], int]:
        header_row, col_desc, col_qty, col_unit = self._find_header(ws)
        if header_row is None or col_desc is None:
            return [], 0

        items: list[ParsedItem] = []
        non_empty = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            desc = row[col_desc] if col_desc < len(row) else None
            desc_s = str(desc).strip() if desc is not None else ""
            if not desc_s:
                continue
            non_empty += 1
            if _is_total_row(desc_s):
                continue

            qty: Decimal | None = None
            unit: str | None = None

            if col_qty is not None and col_qty < len(row):
                qty, unit_from_qty = _parse_qty_with_unit(row[col_qty])
                unit = unit_from_qty

            if col_unit is not None and col_unit < len(row):
                u = row[col_unit]
                if u is not None and str(u).strip():
                    unit = str(u).strip()

            # Если qty не нашли — пробежим по остальным ячейкам в поисках «N <ед.>»
            if qty is None or qty <= 0:
                for j, c in enumerate(row):
                    if j == col_desc or c is None:
                        continue
                    q2, u2 = _parse_qty_with_unit(c)
                    if q2 is not None and q2 > 0:
                        qty = q2
                        if unit is None:
                            unit = u2
                        break

            if qty is None or qty <= 0:
                continue

            items.append(
                ParsedItem(
                    original_text=desc_s,
                    suggested_name=desc_s[:120],
                    suggested_description=desc_s,
                    quantity=qty,
                    unit=(unit or "шт."),
                    raw_data={"sheet": ws.title},
                )
            )
        return items, non_empty

    def _extract_all_text(self, wb: Any) -> str:
        parts: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            parts.append(f"## Лист: {name}")
            for row in ws.iter_rows(values_only=True):
                line = " | ".join("" if c is None else str(c).strip() for c in row)
                if line.strip():
                    parts.append(line)
        return "\n".join(parts)

    def parse_sync(self, file_path: Path) -> tuple[list[ParsedItem], str, float, int]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            all_items: list[ParsedItem] = []
            total_non_empty = 0
            for name in wb.sheetnames:
                ws = wb[name]
                items, non_empty = self._parse_sheet(ws)
                all_items.extend(items)
                total_non_empty += non_empty
            raw = self._extract_all_text(wb)
            if total_non_empty == 0:
                conf = 0.95 if all_items else 0.0
            else:
                ratio = len(all_items) / total_non_empty
                if ratio >= 0.7:
                    conf = 0.95
                elif ratio >= 0.4:
                    conf = 0.6
                else:
                    conf = 0.3
            return all_items, raw, conf, total_non_empty
        finally:
            wb.close()

    async def parse_with_claude(self, raw_text: str) -> ParseResult:
        client = LLMClient()
        prompt = f"""Извлеки ВСЕ позиции из таблицы Excel в JSON.
Файл может содержать несколько листов (заголовки «## Лист: ...»). Каждая строка с описанием — отдельная позиция.
Если количество в виде «27.34 м» / «3 изделия» / «1 шт» — извлекай и число, и единицу.
Если qty не указано — поставь 1.

Формат:
{{"items": [{{"original_text": "...", "suggested_name": "...", "suggested_description": "...", "quantity": число, "unit": "м/шт./м2/..."}}]}}
Только JSON, без markdown.

ТЕКСТ:
{raw_text[:120_000]}"""
        try:
            data = await client.complete_json(
                "Ты извлекаешь строки спецификации из Excel.",
                prompt,
                max_tokens=16000,
                temperature=0,
            )
        except (LLMParseError, Exception) as e:  # noqa: BLE001
            return ParseResult(
                items=[],
                confidence=0.0,
                needs_manual_review=True,
                parser_notes=[str(e)],
            )
        out: list[ParsedItem] = []
        for row in data.get("items") or []:
            q = row.get("quantity")
            try:
                qty = Decimal(str(q)) if q is not None else None
            except InvalidOperation:
                qty = None
            out.append(
                ParsedItem(
                    original_text=str(row.get("original_text", "")),
                    suggested_name=str(row.get("suggested_name", row.get("original_text", ""))),
                    suggested_description=str(row.get("suggested_description", "")),
                    quantity=qty,
                    unit=str(row.get("unit") or "шт."),
                    raw_data={},
                )
            )
        return ParseResult(items=out, confidence=0.7, needs_manual_review=len(out) == 0, parser_notes=[])

    async def parse(self, file_path: Path) -> ParseResult:
        items, raw, conf, _total = await asyncio.to_thread(self.parse_sync, file_path)
        if items and conf >= 0.7:
            return ParseResult(
                items=items,
                confidence=conf,
                needs_manual_review=False,
                parser_notes=[],
                project_metadata={},
            )
        # Структурный парсер уверен слабо или не нашёл ничего → LLM по полному тексту всех листов
        llm_result = await self.parse_with_claude(raw)
        if llm_result.items:
            return llm_result
        if items:
            return ParseResult(
                items=items,
                confidence=conf,
                needs_manual_review=True,
                parser_notes=["Структурный парсер уверен слабо; LLM не дал результата."],
                project_metadata={},
            )
        return llm_result
